import csv
import time
import requests
from bs4 import BeautifulSoup
import utilities.custom_logger as cl
import logging
import openpyxl
from Data.config import Config


class ReadWriteFile:

    log = cl.customLogger(logging.DEBUG)

    def __init__(self):
        self.conf = Config()
        self.excelFile = self.conf.testDataFile

    def getCSVData(fileName):
        # create an empty list to store rows
        rows = []
        # open the CSV file
        dataFile = open(fileName, "r")
        # create a CSV Reader from CSV file
        reader = csv.reader(dataFile)
        # skip the headers
        next(reader)
        # add rows from reader to list
        for row in reader:
            rows.append(row)
        return rows


    def getXMLData(self, url, rootTag, tagToSearch):
        try:
            xmlDict = {}

            r = requests.get(url)
            xmlContent = r.text

            soup = BeautifulSoup(xmlContent, "html.parser")
            sitemapTags = soup.find_all(rootTag)
            self.log.info("The number of {1} tags are {0}".format(len(sitemapTags), rootTag))

            i = 1
            for sitemap in sitemapTags:
                for tag in sitemap.find_all(tagToSearch):
                    xmlDict[i] = tag.text
                    # sitemap.findNext(tagToSearch).text
                    i = i + 1
            self.log.info("##Successfully Traversed File##")
            return xmlDict
        except Exception as e:
            self.log.error(tagToSearch + " Tag not found")
            self.log.error("Exception Occurred " + e)


    def load_excel(self):
        workbook = openpyxl.load_workbook(self.excelFile)
        return workbook


    def read_excel(self, sheetToRead):
        value = []
        workbook = self.load_excel()
        worksheet = workbook[sheetToRead]
        rows = worksheet.max_row
        columns = worksheet.max_column
        for row in rows:
            for column in columns:
                value.append(worksheet.cell(row=row, column=column).value)
        return value


    def count_excelRows(self, sheetName):
        workbook = self.load_excel()
        worksheet = workbook[sheetName]
        rows = worksheet.max_row
        return rows


    def count_excelColumns(self, sheetName):
        workbook = self.load_excel()
        worksheet = workbook[sheetName]
        columns = worksheet.max_column
        return columns


    def write_excel(self, sheetToWrite, row=1, column=1, cellValue=""):
        workbook = self.load_excel()
        worksheet = workbook[sheetToWrite]
        worksheet.cell(row=row, column=column).value = cellValue
        workbook.save(self.excelFile)


    def delete_excelSheet(self, sheetToDelete):
        try:
            workbook = self.load_excel()
            if sheetToDelete in workbook.sheetnames:
                worksheet = workbook[sheetToDelete]
                workbook.remove(worksheet)
                workbook.save(self.excelFile)
                self.log.info("Deleted sheet: " + sheetToDelete)
            else:
                self.log.error("Unable to find the sheet " + sheetToDelete)
        except Exception as e:
            self.log.error("--Exception--Unable to find the sheet " + sheetToDelete + " due to " + str(e))


    def create_excelSheet(self, sheetToCreate):
        workbook = self.load_excel()
        if sheetToCreate not in workbook.sheetnames:
            print(workbook.sheetnames)
            workbook.create_sheet(sheetToCreate, 0)
            workbook.save(self.excelFile)
        else:
            self.log.info("Cannot create sheet ", sheetToCreate, " as it already exists")
        self.log.info("Created sheet " + sheetToCreate)
